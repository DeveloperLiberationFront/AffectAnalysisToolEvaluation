import os
import openpyxl 
import string
path="C://Users/Nasif/Desktop/GitHubdataProject/710results"
os.chdir(path)

#for sentiment
wb=openpyxl.load_workbook('sentimentfinaldata.xlsx')
ws=wb['sentiment']
bucket={}
rates=["positive","negative","neutral"]
tool=["sentiCR"]
tool_col=["M"]
alchemy_error=0
bucket={}
for x in range(0,1):
    for i in rates:
        for j in rates:
            bucket[(i,j)]=0
    for i in range(2,591):
        a=ws['C'+str(i)].value
        if a==None or a=='' or a=='sarcasm':
            #print (i)
            continue
        a=a.strip()
        b=ws[tool_col[x]+str(i)].value
        b=b.strip()
        if(b=="not enough data" or b=="problem"):
            alchemy_error+=1
            continue
        if b=='neg':
            b='negative'
        if b=='pos':
            b='positive'
        bucket[(a,b)]+=1

    if len(bucket)!=9:
        print ("alert! alert! alert!")
    #print (bucket)
    #calculate weighted kohen's kappa

    observation_sum=0

    for k in bucket.keys():
        if k[0]==k[1]:
            observation_sum+=0
            #print (k,bucket[k],observation_sum)
        elif "neutral" in k:
            observation_sum+=bucket[k]
            #print (k,bucket[k],observation_sum)
        else:
            observation_sum=observation_sum+bucket[k]*2
            #print (k,bucket[k],observation_sum)
    #calculate kohen's kappa
    #print (observation_sum)
    agreement=0
    total=0
    values=[]
    for k in bucket.keys():
        total+=bucket[k]
        if k[0] not in  values:
            values.append(k[0])
        if k[1] not in  values:
            values.append(k[1])
        if k[0]==k[1]:
            agreement+=bucket[k]
    #print (agreement,total, values)
    expected_freq={}
    for i in bucket.keys():
        expected_freq[i]=0
        row_total=0
        col_total=0
        for j in bucket.keys():
            if j[0]==i[0]:
                row_total+=bucket[j]
            if j[1]==i[1]:
                col_total+=bucket[j]
        f=(col_total*row_total)/total
        expected_freq[i]=f
    #print (expected_freq)
    expectation_sum=0
    for k in expected_freq.keys():
        if k[0]==k[1]:
            expectation_sum+=0
        elif "neutral" in k:
            expectation_sum+=expected_freq[k]
        else:
            expectation_sum=expectation_sum+expected_freq[k]*2
    weighted_k=1-(observation_sum/expectation_sum)
    print (tool[x], weighted_k)


