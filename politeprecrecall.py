import os
import openpyxl 
import string
path="C://Users/Nasif/Desktop/GitHubdataProject/710results"
os.chdir(path)
wb=openpyxl.load_workbook('sentimentfinaldata.xlsx')
ws=wb['sentiment']
total=0
true_pos=0
false_pos=0
label='neutral'
alchemy_error=0
tool=["sentistrength","NLTK","alchemy","stanfordNLP","senti4SD"]
tool_col=["F","I","G","N","M"]
alchemy_error=0
bucket={}
for x in range(0,5):
    for i in range(2,591):
        a=ws['B'+str(i)].value
        if a==None or a=='' or a=='sarcasm':
            #print (i)
            continue
        a=a.strip()
        b=ws[tool_col[x]+str(i)].value
        b=b.strip()
        if(b=="not enough data" or b=="problem"):
            alchemy_error+=1
            continue
        if b=='neg':
            b='negative'
        if b=='pos':
            b='positive'
        if a==label:
            total+=1
            if b==label:
                true_pos+=1
        else:
            if b==label:
                false_pos+=1
    precision= (true_pos/total)*100
    recall= (true_pos/(true_pos+false_pos))*100
    f_measure= (2*(precision*recall))/(precision+recall)
    print (tool[x],precision,recall,f_measure)
# for i in range(2,591):
#     if ws[col+str(i)].value=="not enough data" or ws[col+str(i)].value=="problem":
#         alchemy_error+=1
#         continue
#     try:
#         if ws['B'+str(i)].value==label:
#             total+=1
#             if ws[col+str(i)].value==label:
#                 true_pos+=1
#         else:
#             if ws[col+str(i)].value==label:
#                 false_pos+=1
#     except:
#         print("problem",i)
print(alchemy_error)