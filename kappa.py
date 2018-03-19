import os
import openpyxl 
path="C://Users/Nasif/Desktop/GitHubdataProject/710results"
os.chdir(path)

#for sentiment
wb=openpyxl.load_workbook('politefinaldata.xlsx')
ws=wb['polite']
bucket={}
total=0
true_pos=0
false_pos=0
label='polite'
for i in range(2,591):
    if ws['C'+str(i)].value==None:
        break
    a=ws['C'+str(i)].value
    if a=='impolite':
        a='neutral'
    if a==None:
        print (i)
    b=ws['G'+str(i)].value
    if (a,b) not in bucket.keys():
        bucket[(a,b)]=1
    else:
        bucket[(a,b)]+=1
    if a==label:
        total+=1
        if b==label:
            true_pos+=1
    else:
        if b==label:
            false_pos+=1
precision= (true_pos/total)*100
recall= (true_pos/(true_pos+false_pos))*100
f_measure= (2*(precision*recall))/(precision+recall)
print (precision,recall,f_measure)
print(i)
print(bucket)
#calculate weighted kohen's kappa
observation_sum=0
for k in bucket.keys():
    if k[0]==k[1]:
        observation_sum+=0
    elif "neutral" in k:
        observation_sum+=bucket[k]
    else:
        observation_sum=observation_sum+bucket[k]*2
#calculate kohen's kappa
agreement=0
total=0
values=[]
for k in bucket.keys():
    total+=bucket[k]
    if k[0] not in  values:
        values.append(k[0])
    if k[1] not in  values:
        values.append(k[1])
    if k[0]==k[1]:
        agreement+=bucket[k]
print (agreement,total, values)
expected_freq={}
for i in bucket.keys():
    expected_freq[i]=0
    row_total=0
    col_total=0
    for j in bucket.keys():
        if j[0]==i[0]:
            row_total+=bucket[j]
        if j[1]==i[1]:
            col_total+=bucket[j]
    f=(col_total*row_total)/total
    expected_freq[i]=f
expectation_sum=0
for k in expected_freq.keys():
    if k[0]==k[1]:
        expectation_sum+=0
    elif "neutral" in k:
        expectation_sum+=expected_freq[k]
    else:
        expectation_sum=expectation_sum+expected_freq[k]*2
weighted_k=1-(observation_sum/expectation_sum)
print (weighted_k)


