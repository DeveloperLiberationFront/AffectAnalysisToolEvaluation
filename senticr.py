import os
import openpyxl 
import string
path="C://Users/Nasif/Desktop/GitHubdataProject/710results"
os.chdir(path)
total=0
true_pos=0
false_pos=0
label='neutral'
#for sentiment
wb=openpyxl.load_workbook('sentiment_firstset_data.xlsx')
ws=wb.active
bucket={}
rates=["negative","neutral","positive"]
tool=["sentiCR"]
tool_col=["H"]
alchemy_error=0
bucket={}
for x in range(0,1):
    for i in rates:
        for j in rates:
            bucket[(i,j)]=0
    for i in range(2,591):
        a=ws['B'+str(i)].value
        if a==None or a=='' or a=='sarcasm':
            #print (i)
            continue
        a=a.strip()
        b=ws[tool_col[x]+str(i)].value
        b=b.strip()
        if(b=="not enough data" or b=="problem"):
            alchemy_error+=1
            continue
        if b=='neg':
            b='negative'
        if b=='pos':
            b='positive'
        bucket[(a,b)]+=1
        if a==label:
            total+=1
            if b==label:
                true_pos+=1
        else:
            if b==label:
                false_pos+=1
    precision= (true_pos/total)*100
    recall= (true_pos/(true_pos+false_pos))*100
    f_measure= (2*(precision*recall))/(precision+recall)
    print (tool[x],precision,recall,f_measure)

    if len(bucket)!=9:
        print ("alert! alert! alert!")
    print(bucket)
    #print (bucket)
    #calculate weighted kohen's kappa

    observation_sum=0

    for k in bucket.keys():
        if k[0]==k[1]:
            observation_sum+=0
            #print (k,bucket[k],observation_sum)
        elif "neutral" in k:
            observation_sum+=bucket[k]
            #print (k,bucket[k],observation_sum)
        else:
            observation_sum=observation_sum+bucket[k]*2
            #print (k,bucket[k],observation_sum)
    #calculate kohen's kappa
    #print (observation_sum)
    observation_sum=0
for k in bucket.keys():
    if k[0]==k[1]:
        observation_sum+=0
    elif "neutral" in k:
        observation_sum+=bucket[k]
    else:
        observation_sum=observation_sum+bucket[k]*2
#calculate kohen's kappa
agreement=0
total=0
values=[]
for k in bucket.keys():
    total+=bucket[k]
    if k[0] not in  values:
        values.append(k[0])
    if k[1] not in  values:
        values.append(k[1])
    if k[0]==k[1]:
        agreement+=bucket[k]
print (agreement,total, values)
expected_freq={}
for i in bucket.keys():
    expected_freq[i]=0
    row_total=0
    col_total=0
    for j in bucket.keys():
        if j[0]==i[0]:
            row_total+=bucket[j]
        if j[1]==i[1]:
            col_total+=bucket[j]
    f=(col_total*row_total)/total
    expected_freq[i]=f
expectation_sum=0
for k in expected_freq.keys():
    if k[0]==k[1]:
        expectation_sum+=0
    elif "neutral" in k:
        expectation_sum+=expected_freq[k]
    else:
        expectation_sum=expectation_sum+expected_freq[k]*2
weighted_k=1-(observation_sum/expectation_sum)
print (weighted_k)