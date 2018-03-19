import os
import openpyxl 
import string
path="C://Users/Nasif/Desktop/GitHubdataProject/710results"
os.chdir(path)

#for sentiment
wb=openpyxl.load_workbook('politefinaldata.xlsx')
ws=wb["polite"]
# total_polite=0
# true_pos=0
# false_pos=0

# total_polite=0
# true_pos=0
# false_pos=0
# for i in range(2,287):
#     if ws['F'+str(i)].value!=None:
#         #print(ws['G'+str(i)].value,ws['F'+str(i)].value)
#         if ws['G'+str(i)].value=='positive':
#                 total_polite+=1
#                 if ws['F'+str(i)].value=="positive":
#                     true_pos+=1
#         else:
#             if ws['F'+str(i)].value=="positive":
#                 false_pos+=1
# #print (true_pos,false_pos)
# precision= (true_pos/total_polite)*100
# recall= (true_pos/(true_pos+false_pos))*100
# f_measure= (2*(precision*recall))/(precision+recall)
# print (true_pos,false_pos, precision,recall,f_measure)
# # #     bar-=.01
# # #print(matched/31,nasif/31)
# # #wb.save('toolsummarygd.xlsx')
    


bucket={}
rates=["polite","neutral"]
#rates=["positive","neutral","negative"]
for i in rates:
    for j in rates:
        bucket[(i,j)]=0
for i in range(2,591):
    a=ws['C'+str(i)].value
    if a==None or a=='':
        print (i)
        continue
    if a=='impolite':
        a='neutral'
    a=a.strip()
    b=ws['G'+str(i)].value
    b=b.strip()   
    bucket[(a,b)]+=1
if len(bucket)!=4:
    print ("alert! alert! alert!")
#print (bucket)
#calculate weighted kohen's kappa

observation_sum=0

for k in bucket.keys():
    if k[0]==k[1]:
        observation_sum+=0
        #print (k,bucket[k],observation_sum)
    elif "neutral" in k:
        observation_sum+=bucket[k]
        #print (k,bucket[k],observation_sum)
    else:
        observation_sum=observation_sum+bucket[k]*2
        #print (k,bucket[k],observation_sum)
#calculate kohen's kappa
#print (observation_sum)
agreement=0
total=0
values=[]
for k in bucket.keys():
    total+=bucket[k]
    if k[0] not in  values:
        values.append(k[0])
    if k[1] not in  values:
        values.append(k[1])
    if k[0]==k[1]:
        agreement+=bucket[k]
#print (agreement,total, values)
expected_freq={}
for i in bucket.keys():
    expected_freq[i]=0
    row_total=0
    col_total=0
    for j in bucket.keys():
        if j[0]==i[0]:
            row_total+=bucket[j]
        if j[1]==i[1]:
            col_total+=bucket[j]
    f=(col_total*row_total)/total
    expected_freq[i]=f
#print (expected_freq)
expectation_sum=0
for k in expected_freq.keys():
    if k[0]==k[1]:
        expectation_sum+=0
    elif "neutral" in k:
        expectation_sum+=expected_freq[k]
    else:
        expectation_sum=expectation_sum+expected_freq[k]*2
weighted_k=1-(observation_sum/expectation_sum)
print(bucket)
print(weighted_k)

    


